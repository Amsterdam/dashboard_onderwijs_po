import logging
import re

from django.db import models
from openpyxl import load_workbook
import pandas

from .vestiging import Vestiging
from .school_wisselaars import _detect_duplicate_brin6, _is_brin6

LOG_FORMAT = '%(asctime)-15s - %(name)s - %(message)s'
logging.basicConfig(format=LOG_FORMAT, level=logging.DEBUG)
logger = logging.getLogger(__name__)


_BRIN6_PATTERN = re.compile('^[0-9]{2}[A-Z]{2}[0-9]{2}$')


def validate_and_split_brin6(value):
    """
    Check that string is brin6, split it into BRIN and vestigingsnummer.
    """
    tmp = value.strip()
    if not _BRIN6_PATTERN.match(tmp):
        msg = 'Value not a valid BRIN6: {}'.format(value)
        raise ValueError(msg)
    return tmp, tmp[:4], tmp[4:]


def _excel_to_dataframe(file_name):
    """
    Convert Excel file with toegewezen subsidies to a Dataframe.
    """
    ws = load_workbook(filename=file_name)['Blad1']
    columns = [x.value for x in ws[4]]
    data = ([x.value for x in row] for row in ws.iter_rows(min_row=5))
    df = pandas.DataFrame(data, columns=columns)

    # Massage columns (rename for consistency, remove useless ones).
    df.rename(
        index=str,
        columns={
            'School brin': 'brin6',
            'School naam': 'school_naam',
        },
        inplace=True
    )
    df.drop('Eindtotaal', axis=1, inplace=True)

    return df


def subsidie_import_helper(file_name, year, brin6s):
    """
    Import ToegewezenSubsidies, creating Subsidies as needed.
    """
    # Assumption: data file contains data for 1 year only.
    # Note: Will fail if run twice.
    df = _excel_to_dataframe(file_name)
    namen = df.columns[2:]  # first two columns are brin6 and school_naam

    # create relevant Subsidie instances, cache them
    _map = {}
    for n in namen:
        s, _ = Subsidie.objects.get_or_create(naam=n, jaar=year)
        _map[n] = s

    # create ToegewezenSubsidie instances
    duplicated_brin6s = _detect_duplicate_brin6(df)

    instances = []
    for i, row in df.iterrows():
        try:
            brin6, brin, vestigingsnummer = validate_and_split_brin6(row['brin6'])
        except ValueError:
            continue
        if brin6 in duplicated_brin6s:
            continue

        for naam, toegewezen in row[2:].iteritems():
            if toegewezen:
                obj = ToegewezenSubsidie(
                    brin=brin,
                    vestigingsnummer=vestigingsnummer,
                    subsidie=_map[naam],
                )
                obj.vestiging_id = brin6 if brin6 in brin6s else None
                instances.append(obj)

    ToegewezenSubsidie.objects.bulk_create(instances)


class Subsidie(models.Model):
    class Meta:
        unique_together = ('jaar', 'naam')

    jaar = models.IntegerField()
    naam = models.CharField(max_length=255)


class ToegewezenSubsidie(models.Model):
    class Meta:
        unique_together = ('subsidie', 'brin', 'vestigingsnummer')

    brin = models.CharField(max_length=4)
    vestigingsnummer = models.IntegerField()

    subsidie = models.ForeignKey(Subsidie)

    vestiging = models.ForeignKey(
        Vestiging, related_name='subsidies', null=True)

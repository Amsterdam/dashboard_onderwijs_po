import logging
import re

from django.db import models
from openpyxl import load_workbook
import pandas

from .vestiging import Vestiging

LOG_FORMAT = '%(asctime)-15s - %(name)s - %(message)s'
logging.basicConfig(format=LOG_FORMAT, level=logging.DEBUG)
logger = logging.getLogger(__name__)

_BRIN6 = re.compile('[0-9][0-9][A-Z][A-Z][0-9][0-9]')


def _detect_duplicate_brin6(dataframe):
    """
    Detect duplicated BRIN 6 codes (with different school names).
    """
    # TODO: proper logging
    mask = dataframe.duplicated(subset=['brin6'])
    duplicated_brin6s = set(dataframe[mask]['brin6'])

    for brin6 in duplicated_brin6s:
        print('Gedupliceerde BRIN6 {} wordt overgeslagen'.format(brin6))
        for i, row in dataframe[dataframe['brin6'] == brin6].iterrows():
            print('  {}: {}'.format(row['brin6'], row['schoolnaam']))

    return duplicated_brin6s


def _is_brin6(value):
    if _BRIN6.match(value):
        return True
    return False


class SchoolWisselaarsManager(models.Manager):
    def _from_excel_file(self, file_name, year, brin6s):
        """
        Load Excel file of "schoolwisselaars", save to db.
        """
        # temp:
        SchoolWisselaars.objects.all().delete()
        # Load the Excel workbook, convert to Pandas DataFrame
        ws = load_workbook(file_name)['schoolwisseling in 2016 perc']
        columns = [x.value for x in ws[1]]
        data = ([x.value for x in row] for row in ws.iter_rows(min_row=2))
        df = pandas.DataFrame(data, columns=columns)

        # Detect / report duplicate BRIN 6 codes.
        duplicated_brin6s = _detect_duplicate_brin6(df)

        # Create instances
        instances = []
        for i, row in df.iterrows():
            if row['brin6'] in duplicated_brin6s:
                continue  # Skip records with duplicated BRIN 6 codes

            brin = row['brin6'][:4]
            vestigingsnummer = int(row['brin6'][4:])

            wisseling_in = row['wisseling in'] if not pandas.isnull(row['wisseling in']) else None
            wisseling_uit = row['wisseling uit'] if not pandas.isnull(row['wisseling uit']) else None

            obj = SchoolWisselaars(
                brin=brin,
                vestigingsnummer=vestigingsnummer,
                naam=row['schoolnaam'],  # potential duplicate of Vestiging naam
                wisseling_uit=wisseling_uit,
                wisseling_in=wisseling_in,
                jaar=year,
            )

            brin6 = row['brin6']
            obj.vestiging_id = brin6 if brin6 in brin6s else None
            instances.append(obj)

        self.bulk_create(instances)


class SchoolWisselaars(models.Model):
    brin = models.CharField(max_length=4)
    vestigingsnummer = models.IntegerField()

    naam = models.CharField(max_length=255)
    wisseling_uit = models.FloatField(null=True)
    wisseling_in = models.FloatField(null=True)

    jaar = models.IntegerField()
    vestiging = models.ForeignKey(
        Vestiging, related_name='schoolwisselaars', null=True)

    objects = SchoolWisselaarsManager()

"""
Models for various external data sources.
"""
# flake8: noqa
# TODO items:
# * consider DecimalField s
import collections
import requests
import pandas
from openpyxl import load_workbook

from datetime import datetime

from django.db import models
from django.core.exceptions import ValidationError

# School types that we save to the database, anything else will fail.
SCHOOL_TYPES = [
    'VMBO_BL',
    'VMBO_BL_KL',
    'VMBO_GT',
    'VMBO_GT_HAVO',
    'VMBO_KL',
    'VMBO_KL_GT',
    'VSO',
    'VWO',
    'HAVO',
    'HAVO_VWO',
    'PRO'
]

# Mapping from school types that we do not store to those we do:
# (There was a slight rename in the data between 2013 and 2014).
SCHOOL_TYPE_MAPPING = collections.OrderedDict([
    ('VMBO_B', 'VMBO_BL'),
    ('VMBO_B_K', 'VMBO_BL_KL'),
    ('VMBO_K', 'VMBO_KL'),
    ('VMBO_K_GT', 'VMBO_KL_GT'),
    ('PRIKDATUM_ADVIEZEN', 'PEILDATUM_ADVIEZEN'),
])

# -- from schoolwijzer --

class Adres(models.Model):
    """Shadow the adress as defined by schoolwijzer.amsterdam.nl"""
    adres = models.CharField(max_length=255)
    email = models.EmailField()
    plaats = models.CharField(max_length=255)
    postcode = models.CharField(max_length=6)
    stadsdeel = models.CharField(max_length=255)
    telefoon = models.CharField(max_length=15)
    website = models.URLField()


class VestigingManager(models.Manager):
    def _from_schoolwijzer_json_enty(self, json_dict):
        pass


class Vestiging(models.Model):
    """Shadow the Vestiging as defined by schoolwijzer.amsterdam.nl"""
    class Meta:
        unique_together = ('brin', 'vestigingsnummer')

    adres = models.ForeignKey(Adres)
    brin = models.CharField(max_length=4)
    lat = models.FloatField()
    lon = models.FloatField()
    grondslag = models.CharField(max_length=255)
    heeft_voorschool = models.BooleanField()
    _id = models.IntegerField()  # TODO: investigate: can be primary key?
    leerlingen = models.IntegerField()
    naam = models.CharField(max_length=255)
    onderwijsconcept = models.CharField(max_length=255, null=True)
    schoolwijzer_url = models.URLField()
    vestigingsnummer = models.IntegerField(null=True)

    brin6 = models.CharField(max_length=6, primary_key=True)
    gebiedscode = models.CharField(max_length=4)


# -- from DUO web APIs and CSV dump

class DUOAPIManagerMixin:
    def from_duo_api_json(self, data, year, brin6s):
        instances = []
        for entry in data['results']:
            instances.append(self._from_duo_api_json_entry(entry, year, brin6s))
        self.bulk_create(instances)


_LEERLING_GEWICHT = {
    'GEWICHT_0': '0',
    'GEWICHT_0.3': '0.3',
    'GEWICHT_1.2': '1.2',
}


class LeerlingNaarGewichtManager(models.Manager):
    def import_csv(self, uri, year, brin6s):
        df = pandas.read_csv(
            uri,
            delimiter=';',
            # dtype=_SCHOOLADVIEZEN_CSV_COLUMNS,  # TODO: add this extra safety
            encoding='cp1252'
        )
        mask = df['GEMEENTENUMMER'] == 363  # only Amsterdam is relevant


        instances = []
        for i, row in df[mask].iterrows():
            brin6 = row['BRIN_NUMMER'] + '{:02d}'.format(int(row['VESTIGINGSNUMMER']))

            # voor ieder gewicht maken we een entry
            for column, value in _LEERLING_GEWICHT.items():  # py3!
                obj = LeerlingNaarGewicht(
                    brin=row['BRIN_NUMMER'],
                    vestigingsnummer=row['VESTIGINGSNUMMER'],

                    gewicht=value,
                    totaal=row[column],
                    jaar=year,
                )

                obj.vestiging_id = brin6 if brin6 in brin6s else None
                instances.append(obj)

        self.bulk_create(instances)


class LeerlingNaarGewicht(models.Model):
    class Meta:
        unique_together = ('brin', 'vestigingsnummer', 'jaar', 'gewicht')

    brin = models.CharField(max_length=4)
    vestigingsnummer = models.IntegerField()

    gewicht = models.CharField(max_length=3)
    totaal = models.IntegerField()
    jaar = models.IntegerField()
    vestiging = models.ForeignKey(Vestiging, null=True, related_name='leerling_naar_gewicht')

    # peildatum = models.DateField()  # TODO: see whether we need peildatum
    objects = LeerlingNaarGewichtManager()


_SCHOOLADVIEZEN_CSV_COLUMNS = collections.OrderedDict([
    ('PEILDATUM_LEERLINGEN', str),
    ('PEILDATUM_ADVIEZEN', str),  # old
    ('PRIKDATUM_ADVIEZEN', str),  # recent
    ('BRIN_NUMMER', str),
    ('VESTIGINGSNUMMER', int),
    ('INSTELLINGSNAAM_VESTIGING', str),
    ('POSTCODE_VESTIGING', str),
    ('PLAATSNAAM', str),
    ('GEMEENTENUMMER', int),
    ('GEMEENTENAAM', str),
    ('PROVINCIE', str),
    ('SOORT_PO', str),
    ('DENOMINATIE_VESTIGING', str),
    ('BEVOEGD_GEZAG_NUMMER', int),

    ('VSO', int),
    ('PRO', int),
    ('VMBO_BL', int),
    ('VMBO_BL_KL', int),
    ('VMBO_KL', int),
    ('VMBO_KL_GT', int),
    ('VMBO_GT', int),
    ('VMBO_GT_HAVO', int),
    ('HAVO', int),
    ('HAVO_VWO', int),
    ('VWO', int),
    ('ADVIES_NIET_MOGELIJK', int),
    ('TOTAAL_ADVIES', int),
])


class SchoolTypeManager(models.Manager):
    def create_known(self):
        '''
        Initialize the SchoolType table, no need to manually create more entries.
        '''
        instances = []
        for name in SCHOOL_TYPES:
            instances.append(SchoolType(name=name))
        self.bulk_create(instances)


class SchoolType(models.Model):
    name = models.CharField(max_length=16, primary_key=True)

    objects = SchoolTypeManager()

    def clean(self):
        if self.name not in SCHOOL_TYPES:
            msg = '{} not a valid schooltype'.format(self.name)
            raise ValidationError(msg)


class SchoolAdviesManager(models.Manager):
    def import_csv(self, uri, year, brin6s):
        """
        Retrieve CSV data from uri create SchoolAdvies entries.
        """
        # TODO: Check encoding='cp1252' with source.
        df = pandas.read_csv(
            uri,
            delimiter=';',
            dtype=_SCHOOLADVIEZEN_CSV_COLUMNS,
            encoding='cp1252'
        )
        mask = df['GEMEENTENUMMER'] == 363  # only Amsterdam is relevant

        # Deal with rename of school types between 2013 and 2014:
        for column_name in df.columns:
            if column_name in SCHOOL_TYPE_MAPPING:
                df[SCHOOL_TYPE_MAPPING[column_name]] = df[column_name]

        # Cretae model instances and save them to database:
        instances = []
        for i, row in df[mask].iterrows():
            brin6 = row['BRIN_NUMMER'] + '{:02d}'.format(int(row['VESTIGINGSNUMMER']))
            peildatum_adviezen =  datetime.strptime(
                str(row['PEILDATUM_ADVIEZEN']), '%Y%m%d')
            peildatum_leerlingen = datetime.strptime(
                row['PEILDATUM_LEERLINGEN'], '%Y%m%d')

            # Make an SchoolAdvies entry for each type of school.
            for school_type in SCHOOL_TYPES:
                obj = SchoolAdvies(
                    brin=row['BRIN_NUMMER'],
                    vestigingsnummer=row['VESTIGINGSNUMMER'],

                    advies_id=school_type,
                    totaal=row[school_type],
                    jaar=year,
                )
                obj.vestiging_id = brin6 if brin6 in brin6s else None
                instances.append(obj)

        self.bulk_create(instances)


class SchoolAdvies(models.Model):
    class Meta:
        unique_together = ('brin', 'vestigingsnummer', 'jaar', 'advies')

    # for internal bookkeeping (keep track of missing vestigingen)
    brin = models.CharField(max_length=4)
    vestigingsnummer = models.IntegerField()

    # to expose to outside
    advies = models.ForeignKey(SchoolType, related_name='school_type')
    totaal = models.IntegerField()
    jaar = models.IntegerField()
    vestiging = models.ForeignKey(Vestiging, null=True, related_name='advies')

    objects = SchoolAdviesManager()


class CitoScoresManager(models.Manager, DUOAPIManagerMixin):
    def _from_duo_api_json_entry(self, json_dict, year, brin6s):
        brin6 = brin=json_dict['BRIN_NUMMER'] + \
            '{:02d}'.format(int(json_dict['VESTIGINGSNUMMER']))

        # Usage of case is not consistent over the years (in the data).
        upper = {key.upper(): key for key in json_dict}

        leerjaar8 = json_dict[upper['LEERJAAR_8']]
        raw = json_dict[upper['CET_GEM']].strip()

        # CITO scores can be missing
        cet_gem = raw.replace(',', '.') if raw else None

        obj = CitoScores(
            brin=json_dict['BRIN_NUMMER'],
            vestigingsnummer=json_dict['VESTIGINGSNUMMER'],
            cet_gem=cet_gem,
            leerjaar_8=leerjaar8,
            jaar=year,
        )
        obj.vestiging_id = brin6 if brin6 in brin6s else None
        return obj


class CitoScores(models.Model):
    """Shadow relevant parst of CITO Scores dataset of DUO."""
    # TODO: peildatum  / prikdatum
    class Meta:
        unique_together = ('brin', 'vestigingsnummer', 'jaar')

    brin = models.CharField(max_length=4)
    vestigingsnummer = models.IntegerField()

    cet_gem = models.FloatField(null=True)
    leerjaar_8 = models.IntegerField(null=True)
    jaar = models.IntegerField()
    vestiging = models.ForeignKey(Vestiging, related_name='cito_scores', null=True)

    objects = CitoScoresManager()

class LeerlingLeraarRatio(models.Model):
    """Model to contain derived number Leerling Leraar Ratio"""
    class Meta:
        unique_together = ('brin', 'jaar')
    # Only BRIN code is present, not vestigingsnummer.
    brin = models.CharField(max_length=4)

    # extracted from Vestigingen table
    n_leerling = models.FloatField(null=True)

    # extracted from DUO CSV file
    n_directie = models.FloatField(null=True)
    n_onderwijzend = models.FloatField(null=True)
    n_ondersteunend = models.FloatField(null=True)
    n_inopleiding = models.FloatField(null=True)
    n_onbekend = models.FloatField(null=True)

    jaar = models.IntegerField()


# -- datasets from onderwijs --

def _detect_duplicate_brin6(dataframe):
    """
    Detect duplicated BRIN 6 codes (with different school names).
    """
    # TODO: proper logging
    mask = dataframe.duplicated(subset=['brin6'])
    duplicated_brin6s = set(dataframe[mask]['brin6'])

    for brin6 in duplicated_brin6s:
        print('Gedupliceerde BRIN6 {} wordt overgeslagen'.format(brin6))
        for i, row in dataframe[dataframe['brin6'] == brin6]:
            print('  {}: {}'.format(row['brin6'], row['schoolnaam']))

    return duplicated_brin6s


class SchoolWisselaarsManager(models.Manager):
    def _from_excel_file(self, file_name, year, brin6s):
        """
        Load Excel file of "schoolwisselaars", save to db.
        """
        # Load the Excel workbook, convert to Pandas DataFrame
        ws = load_workbook(file_name)['schoolwisseling in 2016 perc']
        data = ws.values
        columns = next(data)
        df = DataFrame(data, columns)

        # Detect / report duplicate BRIN 6 codes.
        duplicated_brin6s = _detect_duplicate_brin6(df)

        # Create instances
        instances = []
        for i, row in df.iterrows():
            if row['brin6'] in duplicated_brin6s:
                continue  # Skip records with duplicated BRIN 6 codes

            brin = row['brin6'][:4]
            vestigingsnummer = int(row['brin6'][4:])

            obj = SchoolWisselaars(
                brin=brin,
                vestigingsnummer=vestigingsnummer,
                naam=row['schoolnaam'],  # potential duplicate of Vestiging naam
                wisseling_uit=row['wisseling_uit'],
                wisseling_in=row['wisseling_in'],
                jaar=year,
            )
            obj.vestiging_id = brin6 if brin6 in brin6s else None
            instances.append(obj)

        self.bulk_create(instances)


class SchoolWisselaars(models.Model):
    brin = models.CharField(max_length=4)
    vestigingsnummer = models.IntegerField()

    naam = models.CharField(max_length=255)
    wisseling_uit = models.FloatField()
    wisseling_in = models.FloatField()

    jaar = models.IntegerField()
    vestiging = models.ForeignKey(
        Vestiging, related_name='schoolwisselaars', null=True)


class SubsidieManager(models.Manager):
    def _from_excel(self, file_name, year, brin6s):
        # Load the Excel workbook, convert to Pandas DataFrame
        ws = load_workbook(filename=file_name)['Blad1']
        columns = [x.value for x in ws[4]]
        data = ([x.value for x in row] for row in ws.iter_rows(min_row=5))
        df = pandas.DataFrame(data, columns=columns)
        df.rename({
            'School brin': 'brin6',
            'School naam': 'schoolnaam',
        })  # for consistency with other data sets
        df.drop('Eindtotaal', axis=1, inplace=True)

        duplicated_brin6s = _detect_duplicate_brin6(df)
        instances = []
        for i, row in df.iterrows():
            if row['brin6'] in duplicated_brin6s:
                continue  # Skip records with duplicated BRIN 6 codes

            brin = row['brin6'][:4]
            vestigingsnummer = int(row['brin6'][4:])

            for key, value in row[2:].iteritems():
                if not pandas.isnull(value):
                    # for now only awarded subsidies are stored
                    obj = Subsidie(
                        brin=brin,
                        vestigingsnummer=vestigingsnummer,
                        jaar=year,
                        naam=key,
                    )
                    obj.vestiging_id = brin6 if brin6 in brin6s else None
                    instances.append(obj)

        self.bulk_create(instances)


# The names and available subsidies may vary over the years, so each individual
# subsidy has an entry.
class Subsidie(models.Model):
    brin = models.CharField(max_length=4)
    vestigingsnummer = models.IntegerField()

    jaar = models.IntegerField()
    naam = models.CharField(max_length=255)
    vestiging = models.ForeignKey(
        Vestiging, related_name='subsidies', null=True)

# TODO: VVE data set
